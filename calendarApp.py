#musi byc na poczatku pliku
from __future__ import print_function
import datetime
from googleapiclient.discovery import build
from httplib2 import Http
from oauth2client import file, client, tools

import string

'''
ODCZYTYWANIE PLIKU XLSX Z HARMONOGRAMEM
ORAZ
TOWRZENIE LISTY WSZYSTKICH ZAJEC ZAWARTYCH W ARKUSZU
'''

import openpyxl
from openpyxl import load_workbook

#zmienne globalne - daty rozpoczecia i zakonczenia semestru
data_roz_sob = '2019-02-02T'
data_roz_ndz = '2019-02-03T'
data_zak = '20190623T230000Z'

#zaladowanie harmonogramu 
plan = load_workbook('INF I.xlsx')
ex = plan['6']   #arkusz 6

#funkcja do przetwarzania numerow kolumn na ich odpowiedniki literowe w excelu
def num_to_col_letters(num):
    letters = ''
    while num:
        mod = (num - 1) % 26
        letters += chr(mod + 65)
        num = (num - 1) // 26
    return ''.join(reversed(letters))

#tworzenie listy scalonych komorek
lista = list(ex.merged_cells)
lista2 = list((map(str, lista)))   #czytanie obiektow z listy jako lancuchy tekstowe

#klasa obietkow zajecia do wyeksportowania danych z excela, przystosowany do tworzenia wydarzen w kalendarzu Google
class zajecia:
  def __init__(self, tytul, typ, godz_roz, godz_zak, prowadzacy, sala):
    self.tytul = tytul
    self.typ = typ
    self.godz_roz = godz_roz
    self.godz_zak = godz_zak
    self.prowadzacy = prowadzacy
    self.sala = sala

'''
#funkcja do wypisania wszystkich zajec z biezacego arkusza
def wypisz():
    for i in range(6,ex.max_row+1):
        for j in range(1,ex.max_column+1):
            c = num_to_col_letters(j)
            w = c + str(i) #zapisujemy wspolrzedne sprawdzanej komorki jako lancuch tekstowy
            if ex.cell(row=i, column=j).value!=None:
                #sprawdzanie czy komorka jest scalona
                control = False   #resetujemy zmienna pomocnicza control
                for k in range(0,len(lista2)):
                    control = (w in lista2[k])
                    if control == True: 
                        komscalona = lista2[k]
                        break
                #jesli jest scalona, wylaczamy te, w ktorych sa wpisane godziny    
                if control == True:
                    if(len(komscalona) == 5):
                        if(komscalona[1] == komscalona[4]): continue
                    if(len(komscalona) == 6):
                        print(ex.cell(row=i, column=j).value)
                        print('\n')
                    elif(len(komscalona) == 7):
                        pom1 = komscalona[1:3]
                        pom2 = komscalona[5:7]
                        if(pom1 != pom2):
                            print(ex.cell(row=i, column=j).value)
                            print('\n')

wypisz()
'''

#funkcja pobierajaca dane odnosnie zajec z komorki excela i zwracajaca obiekt klasy zajecia, przystosowany do tworzenia wydarzen w kalendarzu Google
def tworz_zaj(komorka,col,typ):
    zaj = komorka
    zaj = zaj.split('\n')
    if(len(zaj) == 5): 
        zaj[1] = zaj[0] + ' ' + zaj[1]   #Wyklady maja dodatkowy naglowek'WYKŁAD'
        del zaj[0]
    #dostosowanie godziny rozpoczecia i zakonczenia zajec do formatu wymaganego przez kalendarz Google
    t = zaj[2]
    roz, zak = t.split('-')
    zak = zak.split()[0]
    if (len(roz) == 4): roz = '0'+roz
    if (len(zak) == 4): zak = '0'+zak
    if (num_to_col_letters(col) >= 'I' and num_to_col_letters(col) <= 'L'):
        roz, zak = data_roz_sob + roz, data_roz_sob + zak
    if (num_to_col_letters(col) >= 'O' and num_to_col_letters(col) <= 'R'):
        roz, zak = data_roz_ndz + roz, data_roz_ndz + zak
    roz, zak = roz.replace('.',':'), zak.replace('.',':')
    roz, zak = roz+':00',zak+':00'
    zaj1 = zajecia(zaj[0], typ , roz, zak, zaj[1], zaj[3])   #tworzenie obiektu klasy zajecia
    return zaj1

#funkcja zwracajaca typ zajec: Wyklad lub grupe Seminaryjna badz Laboratoryjna
def typ_zajec(zakres, col):
    if(len(zakres) == 7):
        pom1 = ord(zakres[0])
        pom2 = ord(zakres[4])
    else:
        pom1 = ord(zakres[0])
        pom2 = ord(zakres[3])
    if(pom2 - pom1 == 3): typ='Wyklad'
    if(pom2 - pom1 == 1): 
        typ=ex.cell(row=4, column=col).value
    if(pom2 - pom1 == 0):
        typ=ex.cell(row=5, column=col).value
    return typ

#funkcja tworzaca liste wszystkich zajec dostepnych w arkuszu
def lista_zajec():
    lista_zaj = []
    for i in range(6,ex.max_row+1):
        for j in range(1,ex.max_column+1):
            c = num_to_col_letters(j)
            w = c + str(i)   #zapisujemy wspolrzedne sprawdzanej komorki jako lancuch tekstowy
            if ex.cell(row=i, column=j).value!=None:
                #sprawdzanie czy komorka jest scalona
                control = False   #resetujemy zmienna pomocnicza control
                for k in range(0,len(lista2)):
                    control = (w in lista2[k])
                    if control == True: 
                        komscalona = lista2[k]
                        break
                #jesli jest scalona, ignorujemy te, w ktorych sa wpisane godziny (na podtsawie zakresu scalonych komorek)
                if control == True:
                    if(ex.cell(row=i, column=j).value != 'PRZERWA OBIADOWA'):   #ignorujemy przerwe obiadowa
                        if(len(komscalona) == 5):
                            if(komscalona[1] == komscalona[4]): continue
                        if(len(komscalona) == 6):
                            zaj1 = tworz_zaj(ex.cell(row=i, column=j).value,j,typ_zajec(komscalona, j))
                            lista_zaj.append(zaj1)
                        elif(len(komscalona) == 7):
                            pom1 = komscalona[1:3]
                            pom2 = komscalona[5:7]
                            if(pom1 != pom2):
                                zaj1 = tworz_zaj(ex.cell(row=i, column=j).value,j,typ_zajec(komscalona, j))
                                lista_zaj.append(zaj1)
    return lista_zaj

#zapisujemy wszystkie zajecia z biezacego arkusza w liscie wszystkie_zajecia
wszystkie_zajecia = lista_zajec()

#Uzytkownik podaje gr lab i gr sem
Sem = input("Podaj swoją grupę seminaryjną (A1 lub A2):  ")
Lab = input("Podaj swoją grupę laboratoryjną (L1, L2, L3 lub L4):  ")

'''
WPISYWANIE ZAJEC Z WYBRANYCH PRZEZ UZYTKOWNIKA GRUP 
DO ODPOWIEDNICH KALENDARZY GOOGLE
'''

#tworzenie srodowiska pod google api na podstawie kodu udostepnionego w google api developer help
# If modifying these scopes, delete the file token.json.
SCOPES = 'https://www.googleapis.com/auth/calendar'

def main():
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    store = file.Storage('token.json')
    creds = store.get()
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets('credentials.json', SCOPES)
        creds = tools.run_flow(flow, store)
    service = build('calendar', 'v3', http=creds.authorize(Http()))

    # Call the Calendar API
    #now = datetime.datetime.utcnow().isoformat() + 'Z' # 'Z' indicates UTC time

    #tworzenie kalendarzy dla wykladow oraz grupy laboratoryjnej i seminaryjnej wybranej przez uzytkownika
    wyklad = {
        'summary': 'Wykłady',
        'timeZone': 'Europe/Warsaw'
    }
    created_calendar1 = service.calendars().insert(body=wyklad).execute()
    gra = {
        'summary': Sem,
        'timeZone': 'Europe/Warsaw'
    }
    created_calendar2 = service.calendars().insert(body=gra).execute()
    grl = {
        'summary': Lab,
        'timeZone': 'Europe/Warsaw'
    }
    created_calendar3 = service.calendars().insert(body=grl).execute()

    for m in range(len(wszystkie_zajecia)):
        if(wszystkie_zajecia[m].typ == 'Wyklad'):
            #tworzenie eventu typu Wyklad
            event = {
                'summary': wszystkie_zajecia[m].tytul,
                'description': wszystkie_zajecia[m].prowadzacy,
                'location': wszystkie_zajecia[m].sala,
                'start': {
                    'dateTime': wszystkie_zajecia[m].godz_roz,
                    'timeZone': 'Europe/Warsaw',
                },
                'end': {
                    'dateTime': wszystkie_zajecia[m].godz_zak,
                    'timeZone': 'Europe/Warsaw',
                },
                'recurrence': [
                    'RRULE:FREQ=WEEKLY;UNTIL='+data_zak
                ],
            }
            event = service.events().insert(calendarId=created_calendar1['id'], body=event).execute()
        if(wszystkie_zajecia[m].typ == Sem):
            #tworzenie eventu dla wybranej grupy seminaryjnej
            event = {
                'summary': wszystkie_zajecia[m].tytul,
                'description': wszystkie_zajecia[m].prowadzacy,
                'location': wszystkie_zajecia[m].sala,
                'start': {
                    'dateTime': wszystkie_zajecia[m].godz_roz,
                    'timeZone': 'Europe/Warsaw',
                },
                'end': {
                    'dateTime': wszystkie_zajecia[m].godz_zak,
                    'timeZone': 'Europe/Warsaw',
                },
                'recurrence': [
                    'RRULE:FREQ=WEEKLY;UNTIL='+data_zak
                ],
            }
            event = service.events().insert(calendarId=created_calendar2['id'], body=event).execute()
        if(wszystkie_zajecia[m].typ == Lab):
            #tworzenie eventu dla wybranej grupy laboratoryjnej
            event = {
                'summary': wszystkie_zajecia[m].tytul,
                'description': wszystkie_zajecia[m].prowadzacy,
                'location': wszystkie_zajecia[m].sala,
                'start': {
                    'dateTime': wszystkie_zajecia[m].godz_roz,
                    'timeZone': 'Europe/Warsaw',
                },
                'end': {
                    'dateTime': wszystkie_zajecia[m].godz_zak,
                    'timeZone': 'Europe/Warsaw',
                },
                'recurrence': [
                    'RRULE:FREQ=WEEKLY;UNTIL='+data_zak
                ],
            }
            event = service.events().insert(calendarId=created_calendar3['id'], body=event).execute()

    print('Wszystkie zajęcia z harmonogramu dodane do kalendarza Google')


if __name__ == '__main__':
    main()

